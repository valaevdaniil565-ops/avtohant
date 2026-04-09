from __future__ import annotations

import csv
import html
import io
import re
import zipfile
from typing import cast
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from pypdf import PdfReader

from .schemas import NormalizedItem

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("имя", "фио", "кандидат", "специалист"),
    "role": ("роль", "role", "позиция", "вакансия", "должность", "специализация", "специалицация", "title"),
    "stack": ("стек", "стэк", "stack", "технолог", "skills", "skill", "tool", "инструмент"),
    "grade": ("грейд", "уровень", "seniority", "grade"),
    "english": ("англ", "english", "язык"),
    "availability": ("готовность", "доступност", "availability", "start"),
    "location": ("локац", "город", "страна", "регион", "location"),
    "rate": ("ставк", "рейт", "rate", "salary", "cost", "budget"),
    "resume_url": ("резюме", "cv", "resume", "ссылка"),
    "company": ("компан", "company"),
    "client": ("клиент", "заказчик", "client"),
    "work_format": ("формат", "удален", "remote", "office", "hybrid"),
    "employment_type": ("занятост", "оформление", "employment"),
    "requirements": ("требован", "навык", "must have"),
    "responsibilities": ("обязанност", "задач", "responsibilit"),
    "description": ("описание", "комментар", "note", "about"),
}

_GRADE_RE = re.compile(r"(?i)\b(junior|middle\+?|senior|lead|architect|staff|principal|head)\b")
_ENGLISH_RE = re.compile(r"(?i)\b(a1|a2|b1|b2|c1|c2|upper[- ]?intermediate|intermediate|advanced|fluent|native)\b")
_RATE_RE = re.compile(r"(?i)(\d[\d\s]{2,})(?:\s*(₽|руб|rub|\$|usd|eur|€))?")
_ROLEISH_RE = re.compile(
    r"(?i)\b(analyst|designer|developer|engineer|qa|devops|manager|architect|разработ|аналит|дизайн|тест|архитект|менедж)\w*\b"
)
_ABBREV_ROLE_MAP = {
    "са": "Системный аналитик",
    "ca": "Системный аналитик",
    "systemanalyst": "Системный аналитик",
    "системныйаналитик": "Системный аналитик",
    "ба": "Бизнес-аналитик",
    "businessanalyst": "Бизнес-аналитик",
    "бизнесаналитик": "Бизнес-аналитик",
    "qa": "QA Engineer",
    "aqa": "AQA Engineer",
    "pm": "Project Manager",
    "po": "Product Owner",
}
_SKIP_SHEET_NAME_RE = re.compile(
    r"(?i)\b(readme|summary|sum|итог|итоги|свод|archive|архив|template|шаблон|guide|help|инструк|legend|справоч|пример)\b"
)
_TOTALS_RE = re.compile(r"(?i)\b(total|sum|итог|всего|subtotal|результат)\b")
_GENERIC_CONTEXT_WORDS = {
    "bench",
    "vacancy",
    "vacancies",
    "specialists",
    "specialist",
    "candidates",
    "candidate",
    "free",
    "available",
    "свободные",
    "специалисты",
    "специалист",
    "кандидаты",
    "кандидат",
    "вакансии",
    "вакансия",
    "лист",
    "sheet",
}
_STACK_BLOB_LABEL_RE = re.compile(r"(?i)\b(langs?|databases?|devops|stack|tools?|frameworks?)\s*:")


def normalize_table_rows(
    rows: list[list[str]],
    source_url: str,
    *,
    max_items: int = 200,
    table_name: str | None = None,
) -> list[NormalizedItem]:
    items, _ = normalize_table_rows_with_summary(rows, source_url, max_items=max_items, table_name=table_name)
    return items


def normalize_table_rows_with_summary(
    rows: list[list[str]],
    source_url: str,
    *,
    max_items: int = 200,
    table_name: str | None = None,
    sheet_index: int | None = None,
) -> tuple[list[NormalizedItem], dict[str, object]]:
    cleaned = [_trim_row(r) for r in rows]
    nonempty_rows = [r for r in cleaned if any(c.strip() for c in r)]
    summary: dict[str, object] = {
        "sheet_name": table_name or f"Sheet {sheet_index or 1}",
        "sheet_index": sheet_index,
        "sheet_entity_hint": None,
        "sheet_context": {},
        "rows_imported": 0,
        "tables_processed": 0,
        "skip_reasons": {},
        "processed_tables": [],
        "skipped_tables": [],
        "confidence": {"high": 0, "medium": 0, "low": 0},
        "is_skipped": False,
        "skip_reason": None,
    }
    if not nonempty_rows:
        summary["is_skipped"] = True
        summary["skip_reason"] = "empty_sheet"
        return [], summary

    sheet_meta = _classify_sheet(table_name or "", nonempty_rows)
    summary["sheet_entity_hint"] = sheet_meta["entity_hint"]
    summary["sheet_context"] = sheet_meta["context"]
    if not sheet_meta["relevant"]:
        summary["is_skipped"] = True
        summary["skip_reason"] = sheet_meta["reason"]
        return [], summary

    blocks = _split_sheet_into_table_blocks(cleaned)
    if not blocks:
        summary["is_skipped"] = True
        summary["skip_reason"] = "no_detectable_tables"
        return [], summary

    items: list[NormalizedItem] = []
    for block in blocks:
        if len(items) >= max_items:
            break
        block_items, block_summary = _normalize_table_block(
            block,
            source_url,
            sheet_name=table_name,
            sheet_index=sheet_index,
            sheet_meta=sheet_meta,
            max_items=max_items - len(items),
        )
        items.extend(block_items)
        if block_summary.get("rows_imported"):
            summary["tables_processed"] = int(summary["tables_processed"]) + 1
            cast(list, summary["processed_tables"]).append(block_summary)
        else:
            cast(list, summary["skipped_tables"]).append(block_summary)
        summary["rows_imported"] = int(summary["rows_imported"]) + int(block_summary.get("rows_imported") or 0)
        _merge_counter_dict(cast(dict, summary["skip_reasons"]), block_summary.get("skip_reasons") or {})
        _merge_counter_dict(cast(dict, summary["confidence"]), block_summary.get("confidence") or {})

    if not items and not summary["is_skipped"]:
        summary["is_skipped"] = True
        if not summary["skip_reason"]:
            summary["skip_reason"] = "no_confident_rows"
    return items, summary


def csv_bytes_to_items(data: bytes, source_url: str, *, max_items: int = 200) -> list[NormalizedItem]:
    items, _ = csv_bytes_to_items_with_summary(data, source_url, max_items=max_items)
    return items


def csv_bytes_to_items_with_summary(data: bytes, source_url: str, *, max_items: int = 200) -> tuple[list[NormalizedItem], dict[str, object]]:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [[c or "" for c in r] for r in reader]
    items, summary = normalize_table_rows_with_summary(rows, source_url, max_items=max_items)
    workbook_summary = _build_workbook_summary([summary], len(items))
    return items, workbook_summary


def xlsx_bytes_to_items(data: bytes, source_url: str, *, max_items: int = 200) -> list[NormalizedItem]:
    items, _ = xlsx_bytes_to_items_with_summary(data, source_url, max_items=max_items)
    return items


def xlsx_bytes_to_items_with_summary(data: bytes, source_url: str, *, max_items: int = 200) -> tuple[list[NormalizedItem], dict[str, object]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    items: list[NormalizedItem] = []
    sheet_summaries: list[dict[str, object]] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for r in ws.iter_rows(values_only=True):
            rows.append([_cell_to_str(c) for c in r])
        ws_items, sheet_summary = normalize_table_rows_with_summary(
            rows,
            source_url,
            max_items=max_items - len(items),
            table_name=ws.title,
            sheet_index=len(sheet_summaries) + 1,
        )
        sheet_summaries.append(sheet_summary)
        items.extend(ws_items)
        if len(items) >= max_items:
            break
    return items, _build_workbook_summary(sheet_summaries, len(items))


def docx_bytes_to_text(data: bytes) -> str:
    # Minimal parser without python-docx dependency.
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        xml_data = zf.read("word/document.xml")
    root = ET.fromstring(xml_data)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    parts: list[str] = []
    for p in root.findall(".//w:p", ns):
        chunks = []
        for t in p.findall(".//w:t", ns):
            if t.text:
                chunks.append(t.text)
        if chunks:
            parts.append("".join(chunks))
    return "\n".join(parts).strip()


def pdf_bytes_to_text(data: bytes, *, max_pages: int = 30) -> str:
    reader = PdfReader(io.BytesIO(data))
    out: list[str] = []
    for i, p in enumerate(reader.pages):
        if i >= max_pages:
            break
        txt = p.extract_text() or ""
        if txt.strip():
            out.append(txt.strip())
    return "\n\n".join(out).strip()


def html_to_text(html_text: str) -> str:
    # lightweight tag stripper for MVP
    text = re.sub(r"(?is)<script.*?>.*?</script>", " ", html_text)
    text = re.sub(r"(?is)<style.*?>.*?</style>", " ", text)
    text = re.sub(r"(?is)<br\s*/?>", "\n", text)
    text = re.sub(r"(?is)</p>", "\n", text)
    text = re.sub(r"(?is)<[^>]+>", " ", text)
    text = html.unescape(text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _normalize_headers(headers: list[str]) -> list[str]:
    out: list[str] = []
    for i, h in enumerate(headers):
        v = (h or "").strip()
        out.append(v if v else f"Column {i + 1}")
    return out


def _trim_row(row: list[str]) -> list[str]:
    if not row:
        return []
    r = list(row)
    while r and not (r[-1] or "").strip():
        r.pop()
    return r


def _cell_to_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _detect_header_row(rows: list[list[str]]) -> int:
    best_idx = 0
    best_score = float("-inf")
    for idx, row in enumerate(rows[:10]):
        nonempty = [c.strip() for c in row if (c or "").strip()]
        if not nonempty:
            continue
        aliases = [_canonical_header(c) for c in row]
        alias_hits = sum(1 for a in aliases if a)
        unique_hits = len({a for a in aliases if a})
        score = alias_hits * 5 + unique_hits * 2 + len(nonempty)
        if len(nonempty) <= 1:
            score -= 6
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def _header_candidate_metrics(row: list[str]) -> tuple[int, int, int]:
    aliases = [_canonical_header(c) for c in row]
    alias_hits = sum(1 for a in aliases if a)
    unique_hits = len({a for a in aliases if a})
    nonempty = len([c for c in row if (c or "").strip()])
    return alias_hits, unique_hits, nonempty


def _is_header_row_candidate(row: list[str]) -> bool:
    alias_hits, unique_hits, nonempty = _header_candidate_metrics(row)
    return alias_hits >= 2 and unique_hits >= 2 and nonempty >= 2


def _classify_sheet(sheet_name: str, rows: list[list[str]]) -> dict[str, object]:
    label = (sheet_name or "").strip()
    context = _context_hints_from_label(label)
    if label and _SKIP_SHEET_NAME_RE.search(label):
        return {"relevant": False, "reason": "service_sheet_name", "entity_hint": None, "context": context}

    header_rows = [row for row in rows[:25] if _is_header_row_candidate(row)]
    best_header = header_rows[0] if header_rows else None
    entity_hint = None
    if best_header:
        entity_hint = _infer_table_entity_hint([_canonical_header(h) for h in best_header])
    if not entity_hint:
        entity_hint = context.get("entity_hint")

    roleish_rows = 0
    for row in rows[:25]:
        text_value = " ".join((cell or "").strip() for cell in row if (cell or "").strip())
        if text_value and _ROLEISH_RE.search(text_value):
            roleish_rows += 1

    relevant = bool(best_header or roleish_rows >= 2 or context.get("role_hint") or context.get("stack_hints"))
    reason = "sheet_has_headers" if best_header else ("sheet_has_context" if relevant else "no_relevant_structure")
    return {"relevant": relevant, "reason": reason, "entity_hint": entity_hint, "context": context}


def _split_sheet_into_table_blocks(rows: list[list[str]]) -> list[dict[str, object]]:
    header_indices = [idx for idx, row in enumerate(rows) if _is_header_row_candidate(row)]
    if not header_indices and rows:
        candidate = _detect_header_row([r for r in rows if any((c or "").strip() for c in r)])
        nonempty_indices = [idx for idx, row in enumerate(rows) if any((c or "").strip() for c in row)]
        if nonempty_indices:
            header_indices = [nonempty_indices[min(candidate, len(nonempty_indices) - 1)]]

    blocks: list[dict[str, object]] = []
    for pos, header_idx in enumerate(header_indices):
        next_header_idx = header_indices[pos + 1] if pos + 1 < len(header_indices) else len(rows)
        blocks.append(
            {
                "table_index": pos + 1,
                "header_idx": header_idx,
                "rows": rows[header_idx:next_header_idx],
            }
        )
    return blocks


def _normalize_table_block(
    block: dict[str, object],
    source_url: str,
    *,
    sheet_name: str | None,
    sheet_index: int | None,
    sheet_meta: dict[str, object],
    max_items: int,
) -> tuple[list[NormalizedItem], dict[str, object]]:
    table_index = int(block["table_index"])
    header_idx = int(block["header_idx"])
    rows = cast(list[list[str]], block["rows"])
    headers = _normalize_headers(cast(list[str], rows[0]) if rows else [])
    canonical_headers = [_canonical_header(h) for h in headers]
    table_hint = _infer_table_entity_hint(canonical_headers) or cast(str | None, sheet_meta.get("entity_hint"))
    table_summary: dict[str, object] = {
        "table_index": table_index,
        "header_row_index": header_idx + 1,
        "entity_hint": table_hint,
        "rows_imported": 0,
        "skip_reasons": {},
        "confidence": {"high": 0, "medium": 0, "low": 0},
    }
    items: list[NormalizedItem] = []
    section_context: dict[str, object] = {}

    for offset, row in enumerate(rows[1:], start=header_idx + 2):
        if len(items) >= max_items:
            break
        if not any((c or "").strip() for c in row):
            _inc_counter(cast(dict[str, int], table_summary["skip_reasons"]), "empty_row")
            continue
        if _is_totals_row(row):
            _inc_counter(cast(dict[str, int], table_summary["skip_reasons"]), "totals_row")
            continue
        if _is_section_row(row, len(headers)):
            section_context = _context_hints_from_label(_first_nonempty_cell(row))
            _inc_counter(cast(dict[str, int], table_summary["skip_reasons"]), "section_row")
            continue

        row_map: dict[str, str] = {}
        display_lines: list[str] = []
        nonempty = 0
        for i, cell in enumerate(row):
            v = cell.strip()
            if not v:
                continue
            nonempty += 1
            col = headers[i] if i < len(headers) else f"Column {i + 1}"
            canonical = canonical_headers[i] if i < len(canonical_headers) else None
            if canonical and canonical not in row_map:
                row_map[canonical] = v
            display_lines.append(f"{col}: {v}")
        if not display_lines:
            _inc_counter(cast(dict[str, int], table_summary["skip_reasons"]), "empty_row")
            continue
        if len(headers) >= 4 and nonempty <= 1:
            _inc_counter(cast(dict[str, int], table_summary["skip_reasons"]), "section_row")
            continue

        context = _merge_context_hints(cast(dict[str, object], sheet_meta.get("context") or {}), section_context)
        entity_hint = table_hint or _infer_row_entity_hint(canonical_headers, row_map)
        structured, confidence, confidence_reason = _extract_structured_fields(row_map, entity_hint, display_lines, context=context)
        if not structured and entity_hint == "BENCH" and _looks_like_bench_role_only_row(display_lines):
            structured = {"role": display_lines[0].split(":", 1)[-1].strip()}
            confidence = "low"
            confidence_reason = "role_only_row"
        _inc_counter(cast(dict[str, int], table_summary["confidence"]), confidence)
        if confidence == "low":
            _inc_counter(cast(dict[str, int], table_summary["skip_reasons"]), "low_confidence_row")
            continue

        text_lines = _compose_structured_text(display_lines, structured, entity_hint)
        if not text_lines:
            _inc_counter(cast(dict[str, int], table_summary["skip_reasons"]), "no_text_lines")
            continue

        items.append(
            NormalizedItem(
                text="\n".join(text_lines),
                row_index=offset,
                metadata={
                    "source_url": source_url,
                    "table_name": sheet_name,
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "table_index": table_index,
                    "header_row_index": header_idx + 1,
                    "row_map": row_map,
                    "entity_hint": entity_hint,
                    "sheet_entity_hint": sheet_meta.get("entity_hint"),
                    "table_entity_hint": table_hint,
                    "sheet_context": sheet_meta.get("context") or {},
                    "section_context": section_context,
                    "structured_fields": structured,
                    "confidence": confidence,
                    "confidence_reason": confidence_reason,
                },
            )
        )
        table_summary["rows_imported"] = int(table_summary["rows_imported"]) + 1

    if not table_summary["rows_imported"]:
        if not cast(dict[str, int], table_summary["skip_reasons"]):
            _inc_counter(cast(dict[str, int], table_summary["skip_reasons"]), "no_data_rows")
    return items, table_summary


def _context_hints_from_label(label: str) -> dict[str, object]:
    raw = (label or "").strip()
    if not raw:
        return {}
    norm_words = [w for w in re.split(r"[\s/,_\-]+", raw) if w]
    kept_words = [w for w in norm_words if _normalize_token(w) not in _GENERIC_CONTEXT_WORDS]
    cleaned_label = " ".join(kept_words).strip() or raw
    role_hint = _ABBREV_ROLE_MAP.get(_normalize_token(cleaned_label))
    if not role_hint and _ROLEISH_RE.search(cleaned_label):
        role_hint = cleaned_label
    stack_hints = _split_stack_values(cleaned_label)
    availability_hint = None
    low = cleaned_label.lower()
    if "fulltime" in low or "full time" in low:
        availability_hint = "fulltime"
    elif "parttime" in low or "part time" in low:
        availability_hint = "parttime"
    entity_hint = "VACANCY" if re.search(r"(?i)\bvacanc|ваканс", cleaned_label) else ("BENCH" if re.search(r"(?i)\bbench|специал|кандидат", cleaned_label) else None)
    return {
        "label": cleaned_label,
        "role_hint": role_hint,
        "stack_hints": stack_hints,
        "availability_hint": availability_hint,
        "entity_hint": entity_hint,
    }


def _merge_context_hints(base: dict[str, object], extra: dict[str, object]) -> dict[str, object]:
    out = dict(base or {})
    if extra.get("role_hint"):
        out["section_role_hint"] = extra["role_hint"]
    if extra.get("stack_hints"):
        out["section_stack_hints"] = extra["stack_hints"]
    if extra.get("availability_hint"):
        out["availability_hint"] = extra["availability_hint"]
    if extra.get("label"):
        out["section_label"] = extra["label"]
    return out


def _inc_counter(counter: dict[str, int], key: str, by: int = 1) -> None:
    counter[key] = int(counter.get(key) or 0) + by


def _merge_counter_dict(target: dict[str, int], source: dict[str, object]) -> None:
    for key, value in (source or {}).items():
        target[key] = int(target.get(key) or 0) + int(value or 0)


def _build_workbook_summary(sheet_summaries: list[dict[str, object]], items_count: int) -> dict[str, object]:
    processed = [s for s in sheet_summaries if not s.get("is_skipped")]
    skipped = [s for s in sheet_summaries if s.get("is_skipped")]
    confidence = {"high": 0, "medium": 0, "low": 0}
    for summary in processed:
        _merge_counter_dict(confidence, cast(dict[str, object], summary.get("confidence") or {}))
    return {
        "items_count": items_count,
        "sheets_total": len(sheet_summaries),
        "sheets_processed": len(processed),
        "sheets_skipped": len(skipped),
        "processed_sheets": processed,
        "skipped_sheets": skipped,
        "confidence": confidence,
    }


def _is_section_row(row: list[str], headers_len: int) -> bool:
    nonempty = [c.strip() for c in row if (c or "").strip()]
    return headers_len >= 3 and len(nonempty) == 1 and not _TOTALS_RE.search(nonempty[0])


def _is_totals_row(row: list[str]) -> bool:
    first = _first_nonempty_cell(row)
    return bool(first and _TOTALS_RE.search(first))


def _first_nonempty_cell(row: list[str]) -> str:
    for cell in row:
        value = (cell or "").strip()
        if value:
            return value
    return ""


def _canonical_header(header: str) -> str | None:
    norm = re.sub(r"[^a-z0-9а-я]+", " ", (header or "").strip().lower()).strip()
    if not norm:
        return None
    for canonical, aliases in _HEADER_ALIASES.items():
        if any(alias in norm for alias in aliases):
            return canonical
    return None


def _infer_table_entity_hint(canonical_headers: list[str | None]) -> str | None:
    bench_score = 0
    vacancy_score = 0
    header_set = {h for h in canonical_headers if h}
    if "name" in header_set:
        bench_score += 3
    if "availability" in header_set:
        bench_score += 3
    if "resume_url" in header_set:
        bench_score += 2
    if "english" in header_set:
        bench_score += 1
    if "requirements" in header_set:
        vacancy_score += 3
    if "responsibilities" in header_set:
        vacancy_score += 3
    if "company" in header_set or "client" in header_set:
        vacancy_score += 2
    if "employment_type" in header_set or "work_format" in header_set:
        vacancy_score += 1
    if bench_score >= vacancy_score and bench_score >= 3:
        return "BENCH"
    if vacancy_score > bench_score and vacancy_score >= 3:
        return "VACANCY"
    return None


def _infer_row_entity_hint(canonical_headers: list[str | None], row_map: dict[str, str]) -> str | None:
    if "name" in row_map or "availability" in row_map:
        return "BENCH"
    if "requirements" in row_map or "responsibilities" in row_map:
        return "VACANCY"
    if "resume_url" in row_map and "rate" in row_map:
        return "BENCH"
    if any(h in {"company", "client", "employment_type", "work_format"} for h in canonical_headers if h):
        return "VACANCY"
    return None


def _extract_structured_fields(
    row_map: dict[str, str],
    entity_hint: str | None,
    display_lines: list[str],
    *,
    context: dict[str, object] | None = None,
) -> tuple[dict[str, object], str, str]:
    ctx = context or {}
    role = _pick_role(row_map)
    context_role = (
        str(ctx.get("section_role_hint") or "").strip()
        or str(ctx.get("role_hint") or "").strip()
        or None
    )
    if not role and context_role:
        role = context_role

    stack = _pick_stack(row_map, role=role)
    context_stack = cast(list[str], ctx.get("section_stack_hints") or ctx.get("stack_hints") or [])
    if context_stack:
        stack = _dedupe(stack + [v for v in context_stack if v])
    grade = _pick_grade(row_map)
    rate_min, currency = _pick_rate(row_map)
    english = _pick_english(row_map)
    description = "\n".join(display_lines).strip()
    availability = row_map.get("availability") or str(ctx.get("availability_hint") or "").strip() or None
    out: dict[str, object] = {
        "name": row_map.get("name"),
        "role": role,
        "stack": stack,
        "grade": grade,
        "location": row_map.get("location"),
        "availability": availability,
        "rate_min": rate_min,
        "currency": currency,
        "resume_url": row_map.get("resume_url"),
        "english": english,
        "company": row_map.get("company"),
        "client": row_map.get("client"),
        "work_format": row_map.get("work_format"),
        "employment_type": row_map.get("employment_type"),
        "requirements_text": row_map.get("requirements"),
        "responsibilities_text": row_map.get("responsibilities"),
        "description": description,
    }
    if entity_hint == "VACANCY" and not out.get("role"):
        out["role"] = row_map.get("stack")
    cleaned = {k: v for k, v in out.items() if v not in (None, "", [], {})}

    score = 0
    if entity_hint:
        score += 1
    if cleaned.get("role"):
        score += 2
    if cleaned.get("stack"):
        score += 2
    if cleaned.get("grade"):
        score += 1
    if cleaned.get("rate_min") is not None:
        score += 1
    if cleaned.get("location"):
        score += 1
    if cleaned.get("availability"):
        score += 1
    if entity_hint == "BENCH":
        if cleaned.get("name"):
            score += 2
        if cleaned.get("resume_url"):
            score += 1
    if entity_hint == "VACANCY":
        if cleaned.get("company") or cleaned.get("client"):
            score += 1
        if cleaned.get("requirements_text") or cleaned.get("responsibilities_text"):
            score += 2

    used_context = False
    if context_role and cleaned.get("role") == context_role and not row_map.get("role"):
        used_context = True
    if context_stack and not row_map.get("stack"):
        used_context = True
    if availability and not row_map.get("availability") and ctx.get("availability_hint"):
        used_context = True
    if used_context:
        score -= 1

    if score >= 6:
        confidence = "high"
    elif score >= 3:
        confidence = "medium"
    else:
        confidence = "low"

    reason_parts: list[str] = []
    if used_context:
        reason_parts.append("with_context")
    if cleaned.get("name"):
        reason_parts.append("has_name")
    if cleaned.get("role"):
        reason_parts.append("has_role")
    if cleaned.get("stack"):
        reason_parts.append("has_stack")
    if entity_hint == "VACANCY" and (cleaned.get("requirements_text") or cleaned.get("responsibilities_text")):
        reason_parts.append("has_vacancy_details")
    if entity_hint == "BENCH" and cleaned.get("availability"):
        reason_parts.append("has_availability")
    confidence_reason = ",".join(reason_parts) if reason_parts else "weak_row_signal"
    return cleaned, confidence, confidence_reason


def _compose_structured_text(display_lines: list[str], structured: dict[str, object], entity_hint: str | None) -> list[str]:
    if not structured:
        return display_lines
    lines: list[str] = []
    if entity_hint:
        lines.append(f"Entity type hint: {entity_hint}")
    if structured.get("name"):
        lines.append(f"Имя: {structured['name']}")
    if structured.get("role"):
        lines.append(f"Роль: {structured['role']}")
    if structured.get("stack"):
        lines.append(f"Стек: {', '.join(structured['stack'])}")
    if structured.get("grade"):
        lines.append(f"Грейд: {structured['grade']}")
    if structured.get("english"):
        lines.append(f"Английский: {structured['english']}")
    if structured.get("availability"):
        lines.append(f"Готовность: {structured['availability']}")
    if structured.get("location"):
        lines.append(f"Локация: {structured['location']}")
    if structured.get("rate_min") is not None:
        rate = str(structured["rate_min"])
        if structured.get("currency"):
            rate = f"{rate} {structured['currency']}"
        lines.append(f"Ставка: {rate}")
    if structured.get("resume_url"):
        lines.append(f"Резюме: {structured['resume_url']}")
    extra = [ln for ln in display_lines if ln not in lines]
    return lines + extra


def _pick_role(row_map: dict[str, str]) -> str | None:
    direct = row_map.get("role")
    if direct:
        candidate = direct.strip()
        if not _looks_like_stack_blob(candidate) and not _looks_like_urlish(candidate):
            return candidate
    stack_like = (row_map.get("stack") or "").strip()
    if not stack_like:
        return None
    if _looks_like_stack_blob(stack_like) or _looks_like_urlish(stack_like):
        return None
    mapped = _ABBREV_ROLE_MAP.get(_normalize_token(stack_like))
    if mapped:
        return mapped
    if _ROLEISH_RE.search(stack_like) or ("," not in stack_like and "\n" not in stack_like and len(stack_like.split()) <= 5):
        return stack_like
    return None


def _pick_stack(row_map: dict[str, str], *, role: str | None) -> list[str]:
    raw = (row_map.get("stack") or "").strip()
    if not raw:
        return [role] if role else []
    if role and raw == role:
        return [_normalize_role_or_stack(role)]
    parts = _split_stack_values(raw)
    if not parts and role:
        parts = [role]
    if role and len(parts) == 1 and _normalize_token(parts[0]) == _normalize_token(role):
        return [_normalize_role_or_stack(role)]
    return [_normalize_role_or_stack(v) for v in parts if _normalize_role_or_stack(v)]


def _split_stack_values(raw: str) -> list[str]:
    chunks = re.split(r"[\n,;]|(?:\s{2,})", raw)
    out: list[str] = []
    for chunk in chunks:
        v = re.sub(r"(?i)^(langs?|databases?|devops|stack)\s*:\s*", "", chunk.strip())
        if not v:
            continue
        mapped = _ABBREV_ROLE_MAP.get(_normalize_token(v))
        out.append(mapped or v)
    return _dedupe(out)


def _pick_grade(row_map: dict[str, str]) -> str | None:
    for key, value in row_map.items():
        if key == "english":
            continue
        m = _GRADE_RE.search(value)
        if m:
            raw = m.group(1).strip().lower()
            if raw == "middle+":
                return "Middle+"
            return {
                "middle": "Middle",
                "senior": "Senior",
                "junior": "Junior",
                "lead": "Lead",
                "architect": "Architect",
                "head": "Head",
                "staff": "Staff",
                "principal": "Principal",
            }.get(raw, m.group(1))
    return None


def _pick_english(row_map: dict[str, str]) -> str | None:
    value = row_map.get("english")
    if not value:
        return None
    m = _ENGLISH_RE.search(value)
    return m.group(1).upper() if m else value.strip()


def _pick_rate(row_map: dict[str, str]) -> tuple[int | None, str | None]:
    value = (row_map.get("rate") or "").strip()
    if not value:
        return None, None
    m = _RATE_RE.search(value.replace("\xa0", " "))
    if not m:
        return None, None
    amount = int(re.sub(r"\D", "", m.group(1)))
    currency = _normalize_currency(m.group(2) or value)
    return amount, currency


def _normalize_currency(value: str) -> str | None:
    low = (value or "").lower()
    if "usd" in low or "$" in low:
        return "USD"
    if "eur" in low or "€" in low:
        return "EUR"
    if "руб" in low or "rub" in low or "₽" in low:
        return "RUB"
    return None


def _normalize_role_or_stack(value: str) -> str:
    mapped = _ABBREV_ROLE_MAP.get(_normalize_token(value))
    return mapped or value.strip()


def _normalize_token(value: str) -> str:
    return re.sub(r"[^a-z0-9а-я]+", "", (value or "").strip().lower())


def _dedupe(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        key = _normalize_token(value)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(value)
    return out


def _looks_like_bench_role_only_row(display_lines: list[str]) -> bool:
    if len(display_lines) != 1:
        return False
    value = display_lines[0].split(":", 1)[-1].strip()
    return bool(value and _ROLEISH_RE.search(value))


def _looks_like_stack_blob(value: str) -> bool:
    candidate = (value or "").strip()
    if not candidate:
        return False
    if len(candidate) > 120:
        return True
    if "\n" in candidate and _STACK_BLOB_LABEL_RE.search(candidate):
        return True
    if candidate.count(",") >= 5:
        return True
    return False


def _looks_like_urlish(value: str) -> bool:
    candidate = (value or "").strip().lower()
    return bool(candidate and ("http://" in candidate or "https://" in candidate or "docs.google.com/" in candidate or "t.me/" in candidate))
